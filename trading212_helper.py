import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font

# Load your input Excel file
# input_file = 'portfolio_2024.xlsx'
# df = pd.read_excel(input_file)

# Load your input CSV file
input_file = 'portfolio_2024.csv'
df = pd.read_csv(input_file)

# Get unique tickers
tickers = df['Ticker'].unique()

# Create the new Excel file
output_file = 'portfolio_by_ticker.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write the index sheet first
    index_df = pd.DataFrame({'Ticker': tickers})
    index_df.to_excel(writer, sheet_name='Index', index=False)

    # Write each ticker to its own sheet
    for ticker in tickers:
        df_ticker = df[df['Ticker'] == ticker]
        df_ticker.to_excel(writer, sheet_name=ticker, index=False)

# Add hyperlinks using openpyxl
wb = load_workbook(output_file)
index_ws = wb['Index']

# Style header
header_font = Font(bold=True)
index_ws['A1'].font = header_font

# Add hyperlinks to each ticker sheet
for i, ticker in enumerate(tickers, start=2):
    cell = index_ws.cell(row=i, column=1)
    cell.value = ticker
    cell.font = Font(color="0000FF", underline="single")
    cell.hyperlink = f"#{ticker}!A1"  # Hyperlink to A1 cell of the sheet

# Save the modified workbook
wb.save(output_file)